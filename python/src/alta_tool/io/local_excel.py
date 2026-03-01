from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from ..models import OutputRow, PlayerQuery

INPUT_TAB = "Input"
OUTPUT_TAB = "Output"
REQUIRED_INPUT_HEADERS = ["first_name", "last_name"]


class LocalExcelClient:
    def __init__(self, workbook_path: str):
        path = Path(workbook_path).expanduser()
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Local workbook must be an .xlsx file: {path}")
        if not path.exists():
            raise ValueError(f"Local workbook not found: {path}")
        self._workbook_path = path

    def read_input(self) -> list[PlayerQuery]:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(filename=self._workbook_path)
        if INPUT_TAB not in wb.sheetnames:
            raise ValueError(
                "Missing required worksheet tab 'Input'. "
                "Create a tab named exactly 'Input' with headers: "
                "first_name,last_name,city_hint,state_hint"
            )
        ws = wb[INPUT_TAB]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            raise ValueError("Input tab is missing header row")

        header = [_normalize_cell(v) for v in rows[0]]
        if not any(header):
            raise ValueError("Input tab is missing header row")

        col_to_idx = {name: idx for idx, name in enumerate(header) if name}
        missing = [h for h in REQUIRED_INPUT_HEADERS if h not in col_to_idx]
        if missing:
            raise ValueError(f"Input tab missing required columns: {', '.join(missing)}")

        queries: list[PlayerQuery] = []
        for row in rows[1:]:
            first = _row_value(row, col_to_idx, "first_name")
            last = _row_value(row, col_to_idx, "last_name")
            if not first or not last:
                continue
            city = _row_value(row, col_to_idx, "city_hint") or None
            state = (_row_value(row, col_to_idx, "state_hint") or "GA").upper()
            queries.append(
                PlayerQuery(
                    first_name=first,
                    last_name=last,
                    city_hint=city,
                    state_hint=state,
                )
            )
        return queries

    def write_output(self, rows: Iterable[OutputRow]) -> None:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(filename=self._workbook_path)
        if OUTPUT_TAB in wb.sheetnames:
            ws = wb[OUTPUT_TAB]
        else:
            ws = wb.create_sheet(title=OUTPUT_TAB)

        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)

        ws.append(OutputRow.headers())
        notes_col = OutputRow.headers().index("notes") + 1
        ws.cell(row=1, column=notes_col).alignment = openpyxl.styles.Alignment(wrap_text=True)
        for row in rows:
            ws.append(row.to_sheet_row())
            ws.cell(row=ws.max_row, column=notes_col).alignment = openpyxl.styles.Alignment(wrap_text=True)

        wb.save(self._workbook_path)


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_value(row: tuple[Any, ...], col_to_idx: dict[str, int], key: str) -> str:
    idx = col_to_idx.get(key)
    if idx is None:
        return ""
    if idx >= len(row):
        return ""
    return _normalize_cell(row[idx])


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Local Excel mode requires dependency 'openpyxl'.") from exc
    return openpyxl
